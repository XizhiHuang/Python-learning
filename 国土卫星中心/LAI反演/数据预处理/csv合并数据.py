import openpyxl
import os

"""
wb = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")
'2500 0.14449013687747306'
#sheet=wb.get_sheet_names()
sheet=wb.sheetnames
print(sheet)
sheet=wb.get_sheet_by_name('0.3')
a=sheet.cell(1,1)
print(a.value)
print(sheet.dimensions)
a='//lo93r0001_0.3_0_0001.txt'
sheet.cell(3,3).value=a
wb.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")

"""

with open("H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI加密/lo93r0001_0.3_0_0001.txt", "r") as f:  # 打开文件
    data = f.read()  # 读取文件
    print(data)
"""
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/test"):
    tmp_1 = 1
    for file in files:
        # 获取文件的路径及其文件列表
        print(os.path.join(root, file))
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            # with open(data_list, encoding='utf-8').read() as f:
            lines = f.readlines()
            # 筛选出lai文件
            #LAI_value = float(data_list[50:53])
            LAI_value = data_list[44:47]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")
            print(read_data.sheetnames)
            sheet = read_data.get_sheet_by_name(LAI_value)
            print(sheet.dimensions)
            cell_value = sheet.cell(1, 1)
            print(cell_value)
            sheet.cell(1, 1).value = '波长'
            a=sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b=sheet.cell(1, tmp_1)
            i = 1
            n = 2
            for line in lines:
                i = i + 1
                if i > 2:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = tmp_line[0]
                    sheet.cell(n, tmp_1).value = tmp_line[1]
                    n = n + 1
            print("!!!!!!!!!!")
            read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")
print("finish!!!")
"""
# 对每个lai文件夹进行逐个遍历
# LAI=0.3
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_0.3.xlsx")

for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI/0.3"):
    tmp_1 = 1
    for file in files:
        # 获取文件的路径及其文件列表
        # print(os.path.join(root, file))
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            # with open(data_list, encoding='utf-8').read() as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[33:36]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件

            # print(read_data.sheetnames)
            sheet = read_data.get_sheet_by_name(LAI_value)
            # print(sheet.dimensions)
            cell_value = sheet.cell(1, 1)
            # print(cell_value)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            for line in lines:
                i = i + 1
                if i > 2:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = tmp_line[0]
                    sheet.cell(n, tmp_1).value = tmp_line[1]
                    n = n + 1
            # print("!!!!!!!!!!")
            # read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_0.3.xlsx")
print("finish!!!")

# LAI=0.4

read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_0.4.xlsx")

for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI/0.4"):
    tmp_1 = 1
    for file in files:
        # 获取文件的路径及其文件列表
        # print(os.path.join(root, file))
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            # with open(data_list, encoding='utf-8').read() as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[33:36]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件

            # print(read_data.sheetnames)
            sheet = read_data.get_sheet_by_name(LAI_value)
            # print(sheet.dimensions)
            cell_value = sheet.cell(1, 1)
            # print(cell_value)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            for line in lines:
                i = i + 1
                if i > 2:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = tmp_line[0]
                    sheet.cell(n, tmp_1).value = tmp_line[1]
                    n = n + 1
            # print("!!!!!!!!!!")
            # read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_0.4.xlsx")
print("finish!!!")

# LAI=0.5

read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_0.5.xlsx")

for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI/0.5"):
    tmp_1 = 1
    for file in files:
        # 获取文件的路径及其文件列表
        # print(os.path.join(root, file))
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            # with open(data_list, encoding='utf-8').read() as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[33:36]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件

            # print(read_data.sheetnames)
            sheet = read_data.get_sheet_by_name(LAI_value)
            # print(sheet.dimensions)
            cell_value = sheet.cell(1, 1)
            # print(cell_value)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            for line in lines:
                i = i + 1
                if i > 2:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = tmp_line[0]
                    sheet.cell(n, tmp_1).value = tmp_line[1]
                    n = n + 1
            # print("!!!!!!!!!!")
            # read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_0.5.xlsx")
print("finish!!!")

# LAI=0.6

read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_0.6.xlsx")

for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI/0.6"):
    tmp_1 = 1
    for file in files:
        # 获取文件的路径及其文件列表
        # print(os.path.join(root, file))
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            # with open(data_list, encoding='utf-8').read() as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[33:36]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            # read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")
            # print(read_data.sheetnames)
            sheet = read_data.get_sheet_by_name(LAI_value)
            # print(sheet.dimensions)
            cell_value = sheet.cell(1, 1)
            # print(cell_value)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            for line in lines:
                i = i + 1
                if i > 2:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = tmp_line[0]
                    sheet.cell(n, tmp_1).value = tmp_line[1]
                    n = n + 1
            # print("!!!!!!!!!!")
            # read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_0.6.xlsx")
print("finish!!!")

# LAI=0.7

read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_0.7.xlsx")

for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI/0.7"):
    tmp_1 = 1
    for file in files:
        # 获取文件的路径及其文件列表
        print(os.path.join(root, file))
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            # with open(data_list, encoding='utf-8').read() as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[33:36]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            # read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")
            # print(read_data.sheetnames)
            sheet = read_data.get_sheet_by_name(LAI_value)
            # print(sheet.dimensions)
            cell_value = sheet.cell(1, 1)
            # print(cell_value)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            for line in lines:
                i = i + 1
                if i > 2:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = tmp_line[0]
                    sheet.cell(n, tmp_1).value = tmp_line[1]
                    n = n + 1
            # print("!!!!!!!!!!")
            # read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_0.7.xlsx")
print("finish!!!")

# LAI=0.8

read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_0.8.xlsx")

for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI/0.8"):
    tmp_1 = 1
    for file in files:
        # 获取文件的路径及其文件列表
        # print(os.path.join(root, file))
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            # with open(data_list, encoding='utf-8').read() as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[33:36]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            # read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")
            # print(read_data.sheetnames)
            sheet = read_data.get_sheet_by_name(LAI_value)
            # print(sheet.dimensions)
            cell_value = sheet.cell(1, 1)
            # print(cell_value)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            for line in lines:
                i = i + 1
                if i > 2:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = tmp_line[0]
                    sheet.cell(n, tmp_1).value = tmp_line[1]
                    n = n + 1
            # print("!!!!!!!!!!")
            # read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_0.8.xlsx")
print("finish!!!")

# LAI=0.9
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_0.9.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI/0.9"):
    tmp_1 = 1
    for file in files:
        # 获取文件的路径及其文件列表
        # print(os.path.join(root, file))
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            # with open(data_list, encoding='utf-8').read() as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[33:36]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            # read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")
            # print(read_data.sheetnames)
            sheet = read_data.get_sheet_by_name(LAI_value)
            # print(sheet.dimensions)
            cell_value = sheet.cell(1, 1)
            # print(cell_value)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            for line in lines:
                i = i + 1
                if i > 2:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = tmp_line[0]
                    sheet.cell(n, tmp_1).value = tmp_line[1]
                    n = n + 1
            # print("!!!!!!!!!!")
            # read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_0.9.xlsx")
print("finish!!!")

# LAI=1.0

read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_1.0.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI/1.0"):
    tmp_1 = 1
    for file in files:
        # 获取文件的路径及其文件列表
        # print(os.path.join(root, file))
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            # with open(data_list, encoding='utf-8').read() as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[33:36]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            # read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")
            # print(read_data.sheetnames)
            sheet = read_data.get_sheet_by_name(LAI_value)
            # print(sheet.dimensions)
            cell_value = sheet.cell(1, 1)
            # print(cell_value)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            for line in lines:
                i = i + 1
                if i > 2:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = tmp_line[0]
                    sheet.cell(n, tmp_1).value = tmp_line[1]
                    n = n + 1
            # print("!!!!!!!!!!")
            # read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_1.0.xlsx")
print("finish!!!")

# LAI=1.1

read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_1.1.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI/1.1"):
    tmp_1 = 1
    for file in files:
        # 获取文件的路径及其文件列表
        # print(os.path.join(root, file))
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            # with open(data_list, encoding='utf-8').read() as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[33:36]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            # read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")
            # print(read_data.sheetnames)
            sheet = read_data.get_sheet_by_name(LAI_value)
            # print(sheet.dimensions)
            cell_value = sheet.cell(1, 1)
            # print(cell_value)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            for line in lines:
                i = i + 1
                if i > 2:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = tmp_line[0]
                    sheet.cell(n, tmp_1).value = tmp_line[1]
                    n = n + 1
            # print("!!!!!!!!!!")
            # read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_1.1.xlsx")
print("finish!!!")

# LAI=1.2

read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_1.2.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI/1.2"):
    tmp_1 = 1
    for file in files:
        # 获取文件的路径及其文件列表
        # print(os.path.join(root, file))
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            # with open(data_list, encoding='utf-8').read() as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[33:36]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            # read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")
            # print(read_data.sheetnames)
            sheet = read_data.get_sheet_by_name(LAI_value)
            # print(sheet.dimensions)
            cell_value = sheet.cell(1, 1)
            # print(cell_value)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            for line in lines:
                i = i + 1
                if i > 2:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = tmp_line[0]
                    sheet.cell(n, tmp_1).value = tmp_line[1]
                    n = n + 1
            # print("!!!!!!!!!!")
            # read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_1.2.xlsx")
print("finish!!!")

# LAI=1.3
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_1.3.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI/1.3"):
    tmp_1 = 1
    for file in files:
        # 获取文件的路径及其文件列表
        # print(os.path.join(root, file))
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            # with open(data_list, encoding='utf-8').read() as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[33:36]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            # read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")
            # print(read_data.sheetnames)
            sheet = read_data.get_sheet_by_name(LAI_value)
            # print(sheet.dimensions)
            cell_value = sheet.cell(1, 1)
            # print(cell_value)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            for line in lines:
                i = i + 1
                if i > 2:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = tmp_line[0]
                    sheet.cell(n, tmp_1).value = tmp_line[1]
                    n = n + 1
            # print("!!!!!!!!!!")
            # read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_1.3.xlsx")
print("finish!!!")

# LAI=1.4
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_1.4.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI/1.4"):
    tmp_1 = 1
    for file in files:
        # 获取文件的路径及其文件列表
        # print(os.path.join(root, file))
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            # with open(data_list, encoding='utf-8').read() as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[33:36]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            # read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")
            # print(read_data.sheetnames)
            sheet = read_data.get_sheet_by_name(LAI_value)
            # print(sheet.dimensions)
            cell_value = sheet.cell(1, 1)
            # print(cell_value)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            for line in lines:
                i = i + 1
                if i > 2:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = tmp_line[0]
                    sheet.cell(n, tmp_1).value = tmp_line[1]
                    n = n + 1
            # print("!!!!!!!!!!")
            # read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_1.4.xlsx")
print("finish!!!")

# LAI=1.5
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_1.5.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI/1.5"):
    tmp_1 = 1
    for file in files:
        # 获取文件的路径及其文件列表
        # print(os.path.join(root, file))
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            # with open(data_list, encoding='utf-8').read() as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[33:36]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            # read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")
            # print(read_data.sheetnames)
            sheet = read_data.get_sheet_by_name(LAI_value)
            # print(sheet.dimensions)
            cell_value = sheet.cell(1, 1)
            # print(cell_value)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            for line in lines:
                i = i + 1
                if i > 2:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = tmp_line[0]
                    sheet.cell(n, tmp_1).value = tmp_line[1]
                    n = n + 1
            # print("!!!!!!!!!!")
            # read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_1.5.xlsx")
print("finish!!!")

# LAI=1.6
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_1.6.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI/1.6"):
    tmp_1 = 1
    for file in files:
        # 获取文件的路径及其文件列表
        # print(os.path.join(root, file))
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            # with open(data_list, encoding='utf-8').read() as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[33:36]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            # read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")
            # print(read_data.sheetnames)
            sheet = read_data.get_sheet_by_name(LAI_value)
            # print(sheet.dimensions)
            cell_value = sheet.cell(1, 1)
            # print(cell_value)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            for line in lines:
                i = i + 1
                if i > 2:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = tmp_line[0]
                    sheet.cell(n, tmp_1).value = tmp_line[1]
                    n = n + 1
            # print("!!!!!!!!!!")
            # read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_1.6.xlsx")
print("finish!!!")

# LAI=1.7
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_1.7.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI/1.7"):
    tmp_1 = 1
    for file in files:
        # 获取文件的路径及其文件列表
        # print(os.path.join(root, file))
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            # with open(data_list, encoding='utf-8').read() as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[33:36]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            # read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")
            # print(read_data.sheetnames)
            sheet = read_data.get_sheet_by_name(LAI_value)
            # print(sheet.dimensions)
            cell_value = sheet.cell(1, 1)
            # print(cell_value)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            for line in lines:
                i = i + 1
                if i > 2:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = tmp_line[0]
                    sheet.cell(n, tmp_1).value = tmp_line[1]
                    n = n + 1
            # print("!!!!!!!!!!")
            # read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_1.7.xlsx")
print("finish!!!")

# LAI=1.8
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_1.8.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI/1.8"):
    tmp_1 = 1
    for file in files:
        # 获取文件的路径及其文件列表
        # print(os.path.join(root, file))
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            # with open(data_list, encoding='utf-8').read() as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[33:36]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            # read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")
            # print(read_data.sheetnames)
            sheet = read_data.get_sheet_by_name(LAI_value)
            # print(sheet.dimensions)
            cell_value = sheet.cell(1, 1)
            # print(cell_value)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            for line in lines:
                i = i + 1
                if i > 2:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = tmp_line[0]
                    sheet.cell(n, tmp_1).value = tmp_line[1]
                    n = n + 1
            # print("!!!!!!!!!!")
            # read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_1.8.xlsx")
print("finish!!!")

# LAI=1.9
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_1.9.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI/1.9"):
    tmp_1 = 1
    for file in files:
        # 获取文件的路径及其文件列表
        # print(os.path.join(root, file))
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            # with open(data_list, encoding='utf-8').read() as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[33:36]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            # read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")
            # print(read_data.sheetnames)
            sheet = read_data.get_sheet_by_name(LAI_value)
            # print(sheet.dimensions)
            cell_value = sheet.cell(1, 1)
            # print(cell_value)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            for line in lines:
                i = i + 1
                if i > 2:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = tmp_line[0]
                    sheet.cell(n, tmp_1).value = tmp_line[1]
                    n = n + 1
            # print("!!!!!!!!!!")
            # read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_1.9.xlsx")
print("finish!!!")

# LAI=2.0
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_2.0.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI/2.0"):
    tmp_1 = 1
    for file in files:
        # 获取文件的路径及其文件列表
        # print(os.path.join(root, file))
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            # with open(data_list, encoding='utf-8').read() as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[33:36]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            # read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")
            # print(read_data.sheetnames)
            sheet = read_data.get_sheet_by_name(LAI_value)
            # print(sheet.dimensions)
            cell_value = sheet.cell(1, 1)
            # print(cell_value)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            for line in lines:
                i = i + 1
                if i > 2:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = tmp_line[0]
                    sheet.cell(n, tmp_1).value = tmp_line[1]
                    n = n + 1
            # print("!!!!!!!!!!")
            # read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_data.xlsx")

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_2.0.xlsx")
print("finish!!!")
