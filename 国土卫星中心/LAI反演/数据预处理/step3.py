import openpyxl
import os

# LAI=0.3
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample"):
    tmp_1 = 1
    for file in files:
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[52:55]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            sheet = read_data.get_sheet_by_name(LAI_value)
            cell_value = sheet.cell(1, 1)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            band = 1
            for line in lines:
                i = i + 1
                if i > 1:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = band
                    tmp_ref = tmp_line[0]
                    a = tmp_ref[:-1]
                    ref_data = eval(a)
                    sheet.cell(n, tmp_1).value = ref_data
                    n = n + 1
                    band = band + 1

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample.xlsx")
print("finish!!!")
