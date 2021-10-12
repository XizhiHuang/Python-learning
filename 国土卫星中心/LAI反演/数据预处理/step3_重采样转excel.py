import openpyxl
import os

# LAI=0.3
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_0.3.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss/0.3"):
    tmp_1 = 1
    for file in files:
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[48:51]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            sheet = read_data.get_sheet_by_name(LAI_value)
            cell_value = sheet.cell(1, 1)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            band = 1
            for line in lines:
                i = i + 1
                if i > 1:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = band
                    tmp_ref = tmp_line[0]
                    a = tmp_ref[:-1]
                    ref_data = eval(a)
                    sheet.cell(n, tmp_1).value = ref_data
                    n = n + 1
                    band = band + 1

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_0.3.xlsx")
print("finish!!!")

# LAI=0.4
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_0.4.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss/0.4"):
    tmp_1 = 1
    for file in files:
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[48:51]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            sheet = read_data.get_sheet_by_name(LAI_value)
            cell_value = sheet.cell(1, 1)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            band = 1
            for line in lines:
                i = i + 1
                if i > 1:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = band
                    tmp_ref = tmp_line[0]
                    a = tmp_ref[:-1]
                    ref_data = eval(a)
                    sheet.cell(n, tmp_1).value = ref_data
                    n = n + 1
                    band = band + 1

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_0.4.xlsx")
print("finish!!!")

# LAI=0.5
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_0.5.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss/0.5"):
    tmp_1 = 1
    for file in files:
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[48:51]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            sheet = read_data.get_sheet_by_name(LAI_value)
            cell_value = sheet.cell(1, 1)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            band = 1
            for line in lines:
                i = i + 1
                if i > 1:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = band
                    tmp_ref = tmp_line[0]
                    a = tmp_ref[:-1]
                    ref_data = eval(a)
                    sheet.cell(n, tmp_1).value = ref_data
                    n = n + 1
                    band = band + 1

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_0.5.xlsx")
print("finish!!!")

# LAI=0.6
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_0.6.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss/0.6"):
    tmp_1 = 1
    for file in files:
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[48:51]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            sheet = read_data.get_sheet_by_name(LAI_value)
            cell_value = sheet.cell(1, 1)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            band = 1
            for line in lines:
                i = i + 1
                if i > 1:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = band
                    tmp_ref = tmp_line[0]
                    a = tmp_ref[:-1]
                    ref_data = eval(a)
                    sheet.cell(n, tmp_1).value = ref_data
                    n = n + 1
                    band = band + 1

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_0.6.xlsx")
print("finish!!!")

# LAI=0.7
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_0.7.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss/0.7"):
    tmp_1 = 1
    for file in files:
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[48:51]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            sheet = read_data.get_sheet_by_name(LAI_value)
            cell_value = sheet.cell(1, 1)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            band = 1
            for line in lines:
                i = i + 1
                if i > 1:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = band
                    tmp_ref = tmp_line[0]
                    a = tmp_ref[:-1]
                    ref_data = eval(a)
                    sheet.cell(n, tmp_1).value = ref_data
                    n = n + 1
                    band = band + 1

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_0.7.xlsx")
print("finish!!!")

# LAI=0.8
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_0.8.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss/0.8"):
    tmp_1 = 1
    for file in files:
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[48:51]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            sheet = read_data.get_sheet_by_name(LAI_value)
            cell_value = sheet.cell(1, 1)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            band = 1
            for line in lines:
                i = i + 1
                if i > 1:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = band
                    tmp_ref = tmp_line[0]
                    a = tmp_ref[:-1]
                    ref_data = eval(a)
                    sheet.cell(n, tmp_1).value = ref_data
                    n = n + 1
                    band = band + 1

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_0.8.xlsx")
print("finish!!!")

# LAI=0.9
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_0.9.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss/0.9"):
    tmp_1 = 1
    for file in files:
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[48:51]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            sheet = read_data.get_sheet_by_name(LAI_value)
            cell_value = sheet.cell(1, 1)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            band = 1
            for line in lines:
                i = i + 1
                if i > 1:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = band
                    tmp_ref = tmp_line[0]
                    a = tmp_ref[:-1]
                    ref_data = eval(a)
                    sheet.cell(n, tmp_1).value = ref_data
                    n = n + 1
                    band = band + 1

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_0.9.xlsx")
print("finish!!!")

# LAI=1.0
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_1.0.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss/1.0"):
    tmp_1 = 1
    for file in files:
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[48:51]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            sheet = read_data.get_sheet_by_name(LAI_value)
            cell_value = sheet.cell(1, 1)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            band = 1
            for line in lines:
                i = i + 1
                if i > 1:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = band
                    tmp_ref = tmp_line[0]
                    a = tmp_ref[:-1]
                    ref_data = eval(a)
                    sheet.cell(n, tmp_1).value = ref_data
                    n = n + 1
                    band = band + 1

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_1.0.xlsx")
print("finish!!!")

# LAI=1.1
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_1.1.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss/1.1"):
    tmp_1 = 1
    for file in files:
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[48:51]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            sheet = read_data.get_sheet_by_name(LAI_value)
            cell_value = sheet.cell(1, 1)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            band = 1
            for line in lines:
                i = i + 1
                if i > 1:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = band
                    tmp_ref = tmp_line[0]
                    a = tmp_ref[:-1]
                    ref_data = eval(a)
                    sheet.cell(n, tmp_1).value = ref_data
                    n = n + 1
                    band = band + 1

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_1.1.xlsx")
print("finish!!!")

# LAI=1.2
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_1.2.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss/1.2"):
    tmp_1 = 1
    for file in files:
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[48:51]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            sheet = read_data.get_sheet_by_name(LAI_value)
            cell_value = sheet.cell(1, 1)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            band = 1
            for line in lines:
                i = i + 1
                if i > 1:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = band
                    tmp_ref = tmp_line[0]
                    a = tmp_ref[:-1]
                    ref_data = eval(a)
                    sheet.cell(n, tmp_1).value = ref_data
                    n = n + 1
                    band = band + 1

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_1.2.xlsx")
print("finish!!!")

# LAI=1.3
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_1.3.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss/1.3"):
    tmp_1 = 1
    for file in files:
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[48:51]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            sheet = read_data.get_sheet_by_name(LAI_value)
            cell_value = sheet.cell(1, 1)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            band = 1
            for line in lines:
                i = i + 1
                if i > 1:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = band
                    tmp_ref = tmp_line[0]
                    a = tmp_ref[:-1]
                    ref_data = eval(a)
                    sheet.cell(n, tmp_1).value = ref_data
                    n = n + 1
                    band = band + 1

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_1.3.xlsx")
print("finish!!!")

# LAI=1.4
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_1.4.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss/1.4"):
    tmp_1 = 1
    for file in files:
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[48:51]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            sheet = read_data.get_sheet_by_name(LAI_value)
            cell_value = sheet.cell(1, 1)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            band = 1
            for line in lines:
                i = i + 1
                if i > 1:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = band
                    tmp_ref = tmp_line[0]
                    a = tmp_ref[:-1]
                    ref_data = eval(a)
                    sheet.cell(n, tmp_1).value = ref_data
                    n = n + 1
                    band = band + 1

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_1.4.xlsx")
print("finish!!!")

# LAI=1.5
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_1.5.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss/1.5"):
    tmp_1 = 1
    for file in files:
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[48:51]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            sheet = read_data.get_sheet_by_name(LAI_value)
            cell_value = sheet.cell(1, 1)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            band = 1
            for line in lines:
                i = i + 1
                if i > 1:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = band
                    tmp_ref = tmp_line[0]
                    a = tmp_ref[:-1]
                    ref_data = eval(a)
                    sheet.cell(n, tmp_1).value = ref_data
                    n = n + 1
                    band = band + 1

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_1.5.xlsx")
print("finish!!!")

# LAI=1.6
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_1.6.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss/1.6"):
    tmp_1 = 1
    for file in files:
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[48:51]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            sheet = read_data.get_sheet_by_name(LAI_value)
            cell_value = sheet.cell(1, 1)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            band = 1
            for line in lines:
                i = i + 1
                if i > 1:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = band
                    tmp_ref = tmp_line[0]
                    a = tmp_ref[:-1]
                    ref_data = eval(a)
                    sheet.cell(n, tmp_1).value = ref_data
                    n = n + 1
                    band = band + 1

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_1.6.xlsx")
print("finish!!!")

# LAI=1.7
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_1.7.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss/1.7"):
    tmp_1 = 1
    for file in files:
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[48:51]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            sheet = read_data.get_sheet_by_name(LAI_value)
            cell_value = sheet.cell(1, 1)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            band = 1
            for line in lines:
                i = i + 1
                if i > 1:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = band
                    tmp_ref = tmp_line[0]
                    a = tmp_ref[:-1]
                    ref_data = eval(a)
                    sheet.cell(n, tmp_1).value = ref_data
                    n = n + 1
                    band = band + 1

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_1.7.xlsx")
print("finish!!!")

# LAI=1.8
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_1.8.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss/1.8"):
    tmp_1 = 1
    for file in files:
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[48:51]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            sheet = read_data.get_sheet_by_name(LAI_value)
            cell_value = sheet.cell(1, 1)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            band = 1
            for line in lines:
                i = i + 1
                if i > 1:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = band
                    tmp_ref = tmp_line[0]
                    a = tmp_ref[:-1]
                    ref_data = eval(a)
                    sheet.cell(n, tmp_1).value = ref_data
                    n = n + 1
                    band = band + 1

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_1.8.xlsx")
print("finish!!!")

# LAI=1.9
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_1.9.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss/1.9"):
    tmp_1 = 1
    for file in files:
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[48:51]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            sheet = read_data.get_sheet_by_name(LAI_value)
            cell_value = sheet.cell(1, 1)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            band = 1
            for line in lines:
                i = i + 1
                if i > 1:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = band
                    tmp_ref = tmp_line[0]
                    a = tmp_ref[:-1]
                    ref_data = eval(a)
                    sheet.cell(n, tmp_1).value = ref_data
                    n = n + 1
                    band = band + 1

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_1.9.xlsx")
print("finish!!!")

# LAI=2.0
read_data = openpyxl.load_workbook(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_2.0.xlsx")
for root, dirs, files in os.walk(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss/2.0"):
    tmp_1 = 1
    for file in files:
        data_list = os.path.join(root, file)
        with open(data_list, "r")as f:
            lines = f.readlines()
            # 筛选出lai文件
            LAI_value = data_list[48:51]
            LAI_name = data_list[-25:]
            tmp_1 = tmp_1 + 1
            # 打开创建的excel文件
            sheet = read_data.get_sheet_by_name(LAI_value)
            cell_value = sheet.cell(1, 1)
            sheet.cell(1, 1).value = '波长'
            a = sheet.cell(1, 1)
            sheet.cell(1, tmp_1).value = LAI_name
            b = sheet.cell(1, tmp_1)
            i = 1
            n = 2
            band = 1
            for line in lines:
                i = i + 1
                if i > 1:
                    tmp_line = line.split(' ')
                    # 写入data
                    sheet.cell(n, 1).value = band
                    tmp_ref = tmp_line[0]
                    a = tmp_ref[:-1]
                    ref_data = eval(a)
                    sheet.cell(n, tmp_1).value = ref_data
                    n = n + 1
                    band = band + 1

# 保存数据
read_data.save(r"H:/国土卫星中心文档/LAI反演/光谱数据/LAI加密/LAI_resample_gauss_2.0.xlsx")
print("finish!!!")
